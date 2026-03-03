#!/usr/bin/env python3
"""
Statistical comparison of translation method performances using p-values.

Subcommands:
  greek   Within-dataset pairwise + between-dataset (Original vs EIR) for Greek data
  all     Pairwise method comparisons + per-language comparisons for All Languages data

Bonferroni correction applied for multiple comparisons.
"""

import argparse
import itertools
import os

import numpy as np
import pandas as pd
from scipy import stats
from openpyxl.styles import PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)  # parent of the Supplementary Scripts/ directory

EMBEDDING_MODELS = ["cosine_nomic_embed_text", "cosine_sentence_t5_base", "cosine_biolord_2023"]
EMBEDDING_DISPLAY = {
    "cosine_nomic_embed_text": "nomic-embed-text",
    "cosine_sentence_t5_base": "sentence-t5-base",
    "cosine_biolord_2023": "BioLORD-2023",
}

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def load_detailed(path):
    """Load the Detailed sheet from a cross_model_comparison file."""
    return pd.read_excel(path, sheet_name="Detailed")


def color_matrix_sheets(wb, matrix_sheet_names):
    """Apply green/red fills to matrix sheets based on significance (< 0.05)."""
    for name in matrix_sheet_names:
        ws = wb[name]
        for row in ws.iter_rows(min_row=2, min_col=2):  # skip header row/col
            for cell in row:
                if cell.value is None:
                    continue
                try:
                    val = float(cell.value)
                except (ValueError, TypeError):
                    continue
                if val == 1.0:
                    continue  # diagonal
                cell.fill = FILL_GREEN if val < 0.05 else FILL_RED


def build_pvalue_matrix(pairwise_results, label_col=None, label_val=None):
    """Build a symmetric p-value matrix per embedding model for display."""
    df = pd.DataFrame(pairwise_results)
    if label_col and label_val:
        df = df[df[label_col] == label_val]
    sheets = {}
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        sub = df[df["embedding_model"] == emb_name]
        methods = sorted(set(sub["method_1"]) | set(sub["method_2"]))
        matrix = pd.DataFrame(np.nan, index=methods, columns=methods)
        for _, row in sub.iterrows():
            matrix.loc[row["method_1"], row["method_2"]] = row["p_value_bonferroni"]
            matrix.loc[row["method_2"], row["method_1"]] = row["p_value_bonferroni"]
        for m in methods:
            matrix.loc[m, m] = 1.0
        sheets[emb_name] = matrix
    return sheets


# ---------------------------------------------------------------------------
# Greek helpers
# ---------------------------------------------------------------------------

def pairwise_within_dataset(df, label):
    """Compute pairwise paired t-tests between all translation methods for each
    embedding model. Only uses condition_index values shared by both methods.

    Returns a list of result rows.
    """
    methods = sorted(df["translation_method"].unique())
    pairs = list(itertools.combinations(methods, 2))
    n_comparisons = len(pairs)

    results = []
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        for m1, m2 in pairs:
            df1 = df[df["translation_method"] == m1][["condition_index", emb_col]].copy()
            df2 = df[df["translation_method"] == m2][["condition_index", emb_col]].copy()

            # Merge on condition_index for paired test
            merged = df1.merge(df2, on="condition_index", suffixes=("_1", "_2"))
            n = len(merged)

            if n < 2:
                results.append({
                    "dataset": label,
                    "embedding_model": emb_name,
                    "method_1": m1,
                    "method_2": m2,
                    "n_paired": n,
                    "mean_1": np.nan,
                    "mean_2": np.nan,
                    "mean_diff": np.nan,
                    "t_statistic": np.nan,
                    "p_value": np.nan,
                    "p_value_bonferroni": np.nan,
                    "significant_0.05": "",
                })
                continue

            vals1 = merged[f"{emb_col}_1"].values
            vals2 = merged[f"{emb_col}_2"].values

            stat, p = stats.ttest_rel(vals1, vals2)
            p_bonf = min(p * n_comparisons, 1.0)

            results.append({
                "dataset": label,
                "embedding_model": emb_name,
                "method_1": m1,
                "method_2": m2,
                "n_paired": n,
                "mean_1": round(vals1.mean(), 6),
                "mean_2": round(vals2.mean(), 6),
                "mean_diff": round((vals1 - vals2).mean(), 6),
                "t_statistic": round(stat, 6),
                "p_value": round(p, 8),
                "p_value_bonferroni": round(p_bonf, 8),
                "significant_0.05": "Yes" if p_bonf < 0.05 else "No",
            })

    return results


def between_dataset_comparison(df_orig, df_eir):
    """For each translation method and embedding model, independent t-test
    comparing Original vs English Inclusions Removed cosine similarities.

    Returns a list of result rows.
    """
    methods_orig = set(df_orig["translation_method"].unique())
    methods_eir = set(df_eir["translation_method"].unique())
    methods = sorted(methods_orig & methods_eir)
    n_comparisons = len(methods)

    results = []
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        for method in methods:
            orig_vals = df_orig[df_orig["translation_method"] == method][emb_col].values
            eir_vals = df_eir[df_eir["translation_method"] == method][emb_col].values

            n_orig = len(orig_vals)
            n_eir = len(eir_vals)
            n_removed = n_orig - n_eir

            if n_orig < 2 or n_eir < 2:
                results.append({
                    "embedding_model": emb_name,
                    "translation_method": method,
                    "n_original": n_orig,
                    "n_eir": n_eir,
                    "n_removed": n_removed,
                    "mean_original": np.nan,
                    "std_original": np.nan,
                    "mean_eir": np.nan,
                    "std_eir": np.nan,
                    "mean_diff": np.nan,
                    "t_statistic": np.nan,
                    "p_value": np.nan,
                    "p_value_bonferroni": np.nan,
                    "significant_0.05": "",
                })
                continue

            stat, p = stats.ttest_ind(orig_vals, eir_vals)
            p_bonf = min(p * n_comparisons, 1.0)

            results.append({
                "embedding_model": emb_name,
                "translation_method": method,
                "n_original": n_orig,
                "n_eir": n_eir,
                "n_removed": n_removed,
                "mean_original": round(orig_vals.mean(), 6),
                "std_original": round(orig_vals.std(ddof=1), 6),
                "mean_eir": round(eir_vals.mean(), 6),
                "std_eir": round(eir_vals.std(ddof=1), 6),
                "mean_diff": round(eir_vals.mean() - orig_vals.mean(), 6),
                "t_statistic": round(stat, 6),
                "p_value": round(p, 8),
                "p_value_bonferroni": round(p_bonf, 8),
                "significant_0.05": "Yes" if p_bonf < 0.05 else "No",
            })

    return results


# ===========================================================================
# Subcommand: greek
# ===========================================================================

def run_greek():
    """Within-dataset pairwise + between-dataset (Original vs EIR) for Greek data."""
    greek_dir = os.path.join(PROJECT_DIR, "Greek Data")
    original_dir = os.path.join(greek_dir, "Greek Data (Original)", "Semantic Value Comparisons")
    eir_dir = os.path.join(greek_dir, "Greek Data (English Inclusions Removed)", "Semantic Value Comparisons")

    print("Loading cross-model comparison data...")
    orig_path = os.path.join(original_dir, "cross_model_comparison.xlsx")
    eir_path = os.path.join(eir_dir, "cross_model_comparison.xlsx")

    df_orig = load_detailed(orig_path)
    df_eir = load_detailed(eir_path)
    print(f"  Original: {len(df_orig)} rows")
    print(f"  English Inclusions Removed: {len(df_eir)} rows")

    # ==================================================================
    # 1. Within-dataset pairwise comparisons
    # ==================================================================
    print(f"\n{'='*60}")
    print("WITHIN-DATASET PAIRWISE COMPARISONS (paired t-tests)")
    print(f"{'='*60}")

    print("\n  Computing Original dataset pairwise tests...")
    orig_pairwise = pairwise_within_dataset(df_orig, "Original")
    print(f"    {len(orig_pairwise)} comparisons")

    print("  Computing English Inclusions Removed pairwise tests...")
    eir_pairwise = pairwise_within_dataset(df_eir, "English Inclusions Removed")
    print(f"    {len(eir_pairwise)} comparisons")

    # Write Original p-values
    orig_out = os.path.join(original_dir, "p_value_analysis.xlsx")
    orig_df = pd.DataFrame(orig_pairwise)
    orig_matrices = build_pvalue_matrix(orig_pairwise, "dataset", "Original")

    orig_matrix_sheets = []
    with pd.ExcelWriter(orig_out, engine="openpyxl") as writer:
        orig_df.drop(columns=["dataset"]).to_excel(writer, sheet_name="Pairwise Detail", index=False)
        for emb_name, matrix in orig_matrices.items():
            safe_name = emb_name.replace("/", "_")[:31]
            sheet_name = f"Matrix {safe_name}"
            matrix.to_excel(writer, sheet_name=sheet_name)
            orig_matrix_sheets.append(sheet_name)
        color_matrix_sheets(writer.book, orig_matrix_sheets)

    print(f"\n  Original p-values: {orig_out}")

    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        sub = orig_df[orig_df["embedding_model"] == emb_name]
        sig = sub[sub["significant_0.05"] == "Yes"]
        print(f"    {emb_name}: {len(sig)}/{len(sub)} significant pairs (Bonferroni p < 0.05)")

    # Write EIR p-values
    eir_out = os.path.join(eir_dir, "p_value_analysis.xlsx")
    eir_df = pd.DataFrame(eir_pairwise)
    eir_matrices = build_pvalue_matrix(eir_pairwise, "dataset", "English Inclusions Removed")

    eir_matrix_sheets = []
    with pd.ExcelWriter(eir_out, engine="openpyxl") as writer:
        eir_df.drop(columns=["dataset"]).to_excel(writer, sheet_name="Pairwise Detail", index=False)
        for emb_name, matrix in eir_matrices.items():
            safe_name = emb_name.replace("/", "_")[:31]
            sheet_name = f"Matrix {safe_name}"
            matrix.to_excel(writer, sheet_name=sheet_name)
            eir_matrix_sheets.append(sheet_name)
        color_matrix_sheets(writer.book, eir_matrix_sheets)

    print(f"  EIR p-values: {eir_out}")

    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        sub = eir_df[eir_df["embedding_model"] == emb_name]
        sig = sub[sub["significant_0.05"] == "Yes"]
        print(f"    {emb_name}: {len(sig)}/{len(sub)} significant pairs (Bonferroni p < 0.05)")

    # ==================================================================
    # 2. Between-dataset comparison (Original vs EIR)
    # ==================================================================
    print(f"\n{'='*60}")
    print("BETWEEN-DATASET COMPARISON (independent t-tests)")
    print(f"{'='*60}")

    between_results = between_dataset_comparison(df_orig, df_eir)
    between_df = pd.DataFrame(between_results)

    between_out = os.path.join(greek_dir, "between_dataset_p_value_analysis.xlsx")
    between_df.to_excel(between_out, index=False)
    print(f"\n  Between-dataset p-values: {between_out}")

    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        sub = between_df[between_df["embedding_model"] == emb_name]
        print(f"\n  {emb_name}:")
        for _, row in sub.iterrows():
            sig_marker = " *" if row["significant_0.05"] == "Yes" else ""
            print(f"    {row['translation_method']:30s}  orig={row['mean_original']:.4f}  eir={row['mean_eir']:.4f}  "
                  f"diff={row['mean_diff']:+.4f}  p_bonf={row['p_value_bonferroni']:.6f}{sig_marker}")

    print(f"\n{'='*60}")
    print("DONE!")
    print(f"{'='*60}")


# ===========================================================================
# Subcommand: all
# ===========================================================================

def run_all():
    """Pairwise method comparisons + per-language comparisons for All Languages data."""
    svc_dir = os.path.join(PROJECT_DIR, "Semantic Value Comparisons (All Languages)")

    cross_path = os.path.join(svc_dir, "cross_model_comparison.xlsx")
    print(f"Loading: {cross_path}")
    df = pd.read_excel(cross_path, sheet_name="Detailed")
    print(f"  {len(df)} rows loaded.")

    methods = sorted(df["method"].unique())
    pairs = list(itertools.combinations(methods, 2))
    n_comparisons = len(pairs)
    print(f"  Methods: {methods}")
    print(f"  Pairwise comparisons: {n_comparisons}")

    # --- Pairwise by method (paired on condition_index + language) ---
    print(f"\n{'='*60}")
    print("PAIRWISE METHOD COMPARISONS (paired t-tests)")
    print(f"{'='*60}")

    pairwise_results = []
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        for m1, m2 in pairs:
            df1 = df[df["method"] == m1][["condition_index", "target_language_code", emb_col]].copy()
            df2 = df[df["method"] == m2][["condition_index", "target_language_code", emb_col]].copy()

            merged = df1.merge(df2, on=["condition_index", "target_language_code"], suffixes=("_1", "_2"))
            n = len(merged)

            if n < 2:
                pairwise_results.append({
                    "embedding_model": emb_name,
                    "method_1": m1,
                    "method_2": m2,
                    "n_paired": n,
                    "mean_1": np.nan,
                    "mean_2": np.nan,
                    "mean_diff": np.nan,
                    "t_statistic": np.nan,
                    "p_value": np.nan,
                    "p_value_bonferroni": np.nan,
                    "significant_0.05": "",
                })
                continue

            vals1 = merged[f"{emb_col}_1"].values
            vals2 = merged[f"{emb_col}_2"].values

            if np.all(vals1 == vals2):
                stat, p, p_bonf = 0.0, 1.0, 1.0
            else:
                stat, p = stats.ttest_rel(vals1, vals2)
                p_bonf = min(p * n_comparisons, 1.0)

            pairwise_results.append({
                "embedding_model": emb_name,
                "method_1": m1,
                "method_2": m2,
                "n_paired": n,
                "mean_1": round(vals1.mean(), 6),
                "mean_2": round(vals2.mean(), 6),
                "mean_diff": round((vals1 - vals2).mean(), 6),
                "t_statistic": round(stat, 6),
                "p_value": round(p, 8),
                "p_value_bonferroni": round(p_bonf, 8),
                "significant_0.05": "Yes" if p_bonf < 0.05 else "No",
            })

            print(f"  {emb_name}: {m1} vs {m2}  (n={n})  "
                  f"mean_diff={vals1.mean()-vals2.mean():+.6f}  "
                  f"p_bonf={p_bonf:.8f}  "
                  f"{'*' if p_bonf < 0.05 else ''}")

    pairwise_df = pd.DataFrame(pairwise_results)

    # --- Pairwise by method PER LANGUAGE ---
    print(f"\n{'='*60}")
    print("PAIRWISE METHOD COMPARISONS BY LANGUAGE (paired t-tests)")
    print(f"{'='*60}")

    languages = sorted(df["target_language_code"].unique())
    lang_name_map = dict(zip(df["target_language_code"], df["target_language_name"]))

    lang_results = []
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        for lang in languages:
            df_lang = df[df["target_language_code"] == lang]
            lang_methods = sorted(df_lang["method"].unique())
            lang_pairs = list(itertools.combinations(lang_methods, 2))
            n_lang_comparisons = len(lang_pairs)

            for m1, m2 in lang_pairs:
                df1 = df_lang[df_lang["method"] == m1][["condition_index", emb_col]].copy()
                df2 = df_lang[df_lang["method"] == m2][["condition_index", emb_col]].copy()

                merged = df1.merge(df2, on="condition_index", suffixes=("_1", "_2"))
                n = len(merged)

                if n < 2:
                    continue

                vals1 = merged[f"{emb_col}_1"].values
                vals2 = merged[f"{emb_col}_2"].values

                if np.all(vals1 == vals2):
                    stat, p, p_bonf = 0.0, 1.0, 1.0
                else:
                    stat, p = stats.ttest_rel(vals1, vals2)
                    p_bonf = min(p * n_lang_comparisons, 1.0)

                lang_results.append({
                    "embedding_model": emb_name,
                    "target_language_code": lang,
                    "target_language_name": lang_name_map.get(lang, lang),
                    "method_1": m1,
                    "method_2": m2,
                    "n_paired": n,
                    "mean_1": round(vals1.mean(), 6),
                    "mean_2": round(vals2.mean(), 6),
                    "mean_diff": round((vals1 - vals2).mean(), 6),
                    "t_statistic": round(stat, 6),
                    "p_value": round(p, 8),
                    "p_value_bonferroni": round(p_bonf, 8),
                    "significant_0.05": "Yes" if p_bonf < 0.05 else "No",
                })

    lang_df = pd.DataFrame(lang_results)

    # Console summary
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        sub = lang_df[lang_df["embedding_model"] == emb_name]
        sig = sub[sub["significant_0.05"] == "Yes"]
        total = len(sub)
        print(f"  {emb_name}: {len(sig)}/{total} significant language-method pairs")

    # --- Build p-value matrices (method x method) ---
    matrix_sheets = {}
    for emb_col in EMBEDDING_MODELS:
        emb_name = EMBEDDING_DISPLAY[emb_col]
        sub = pairwise_df[pairwise_df["embedding_model"] == emb_name]
        matrix = pd.DataFrame(np.nan, index=methods, columns=methods)
        for _, row in sub.iterrows():
            matrix.loc[row["method_1"], row["method_2"]] = row["p_value_bonferroni"]
            matrix.loc[row["method_2"], row["method_1"]] = row["p_value_bonferroni"]
        for m in methods:
            matrix.loc[m, m] = 1.0
        matrix_sheets[emb_name] = matrix

    # --- Write output ---
    out_path = os.path.join(svc_dir, "p_value_analysis.xlsx")

    matrix_sheet_names = []
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pairwise_df.to_excel(writer, sheet_name="Pairwise Detail", index=False)

        for emb_name, matrix in matrix_sheets.items():
            safe_name = emb_name.replace("/", "_")[:31]
            sheet_name = f"Matrix {safe_name}"
            matrix.to_excel(writer, sheet_name=sheet_name)
            matrix_sheet_names.append(sheet_name)

        lang_df.to_excel(writer, sheet_name="By Language Detail", index=False)

        color_matrix_sheets(writer.book, matrix_sheet_names)

    print(f"\n  Results: {out_path}")
    print(f"\n{'='*60}")
    print("DONE!")
    print(f"{'='*60}")


# ===========================================================================
# Main
# ===========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Statistical comparison of translation method performances using p-values.",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    subparsers.add_parser("greek", help="Within-dataset pairwise + between-dataset (Original vs EIR)")
    subparsers.add_parser("all", help="Pairwise method comparisons + per-language for All Languages")

    args = parser.parse_args()

    if args.command == "greek":
        run_greek()
    elif args.command == "all":
        run_all()


if __name__ == "__main__":
    main()
